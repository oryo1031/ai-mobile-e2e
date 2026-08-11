"""設計書の取り込み。

設計書は実際には Excel や Confluence にあり、それを手で Markdown に起こす作業が
このワークフローの一番の負担になっていた。ここはその変換を機械化する。

変換の目的は「人が読んで美しい Markdown」を作ることではなく、後続のエージェントが
情報を落とさずに読めることにある。体裁より情報の保存を優先する。多少雑でも安全に
倒れるのは、仕様の正規化工程に open_questions のゲートがあるため。変換で情報が
欠けていれば、spec-analyst が推測で埋めずに質問として差し戻す。
"""

from __future__ import annotations

import csv as csv_module
import datetime
import io
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".html", ".htm", ".csv", ".md", ".txt"}

#: 表がこの列数を超えたら、空列を落としても読みづらいので警告する。
WIDE_TABLE_COLUMNS = 20


class IngestError(RuntimeError):
    """取り込めない入力。"""


@dataclass
class IngestResult:
    output: Path
    source: Path
    sections: int
    warnings: list[str]


# ----------------------------------------------------------------------
# 共通
# ----------------------------------------------------------------------
def _escape_cell(value: object) -> str:
    """Markdown の表に入れられる形にする。"""
    if value is None:
        return ""
    text = str(value).strip()
    # 表を壊さないよう、パイプと改行を退避する。
    return text.replace("|", "\\|").replace("\r\n", " / ").replace("\n", " / ")


def _render_table(rows: list[list[str]]) -> list[str]:
    """行の並びを Markdown の表にする。

    設計書の表は 1 行目が見出しとは限らないが、Markdown の表は見出し行を
    要求する。1 行目を見出しとして扱い、そうでない場合も情報は失われない。

    先頭が横結合されたタイトル行のときは、全列に同じ値が展開されて
    見出しが埋まってしまうため、表から外して 1 行の見出しとして出す。
    """
    if not rows:
        return []
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]

    lead: list[str] = []
    while (
        len(padded) > 1
        and width > 1
        and padded[0][0]
        and all(cell == padded[0][0] for cell in padded[0])
    ):
        lead.extend([f"**{padded[0][0]}**", ""])
        padded = padded[1:]
        padded = _trim_empty(padded) or padded
        width = max(len(row) for row in padded)
        padded = [row + [""] * (width - len(row)) for row in padded]

    lines = [*lead, "| " + " | ".join(padded[0]) + " |", "|" + "---|" * width]
    for row in padded[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return lines


def _trim_empty(rows: list[list[str]]) -> list[list[str]]:
    """全体が空の行と列を落とす。"""
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        return []
    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    keep = [i for i in range(width) if any(row[i] for row in padded)]
    return [[row[i] for i in keep] for row in padded]


# ----------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------
def _read_xlsx(path: Path, sheets: list[str] | None) -> tuple[list[str], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 依存が入っていない環境向け
        raise IngestError(
            "Excel の取り込みには openpyxl が必要です。`uv sync` を実行してください。"
        ) from exc

    # data_only=True で数式ではなく計算済みの値を読む。
    workbook = load_workbook(path, data_only=True, read_only=False)
    lines: list[str] = []
    warnings: list[str] = []
    count = 0

    for sheet in workbook.worksheets:
        if sheets and sheet.title not in sheets:
            continue
        if sheet.sheet_state != "visible":
            warnings.append(f"シート『{sheet.title}』は非表示のため取り込みません。")
            continue

        grid: list[list[str]] = [
            [_escape_cell(cell.value) for cell in row] for row in sheet.iter_rows()
        ]
        if not grid:
            continue

        # 結合セルの値を被覆範囲へ展開する。日本語の設計書は結合が多用され、
        # これをやらないと項目名が消えて表が読めなくなる。
        for merged in sheet.merged_cells.ranges:
            top, left = merged.min_row - 1, merged.min_col - 1
            if top >= len(grid) or left >= len(grid[top]):
                continue
            value = grid[top][left]
            if not value:
                continue
            for r in range(merged.min_row - 1, merged.max_row):
                for c in range(merged.min_col - 1, merged.max_col):
                    if r < len(grid) and c < len(grid[r]):
                        grid[r][c] = value

        trimmed = _trim_empty(grid)
        if not trimmed:
            continue
        if len(trimmed[0]) > WIDE_TABLE_COLUMNS:
            warnings.append(
                f"シート『{sheet.title}』は {len(trimmed[0])} 列あります。"
                "レイアウト用のセルが混ざっていないか確認してください。"
            )

        count += 1
        lines.append(f"## シート: {sheet.title}")
        lines.append("")
        lines.extend(_render_table(trimmed))
        lines.append("")

    workbook.close()
    if count == 0:
        warnings.append("取り込める内容のあるシートがありませんでした。")
    return lines, warnings


# ----------------------------------------------------------------------
# HTML (Confluence のページエクスポート)
# ----------------------------------------------------------------------
class _HtmlToMarkdown(HTMLParser):
    """Confluence のエクスポート HTML から本文を取り出す。

    体裁の再現は狙わず、見出し・段落・箇条書き・表だけを拾う。

    エクスポートした HTML には、パンくず・サイドバー・コメント・ラベルといった
    Confluence の画面部品が本文と一緒に入ってくる。そのまま取り込むと後続の
    エージェントが仕様と見分けられないため、本文の領域を特定して
    そこだけを採用する。
    """

    SKIP_TAGS = {"script", "style", "head", "nav", "footer", "noscript"}
    HEADINGS = {"h1", "h2", "h3", "h4", "h5", "h6"}

    #: 本文が入っている入れ物。Confluence の Server/DC と Cloud の両方を見る。
    CONTENT_IDS = {"main-content", "content"}
    CONTENT_CLASSES = {"wiki-content", "ak-renderer-document", "page-content"}

    #: 本文ではない画面部品。id / class の部分一致で落とす。
    CHROME_MARKERS = (
        "breadcrumb",
        "comment",
        "sidebar",
        "page-metadata",
        "labels-section",
        "likes-and-labels",
        "aui-nav",
        "footer",
        "page-restrictions",
        "inline-comment",
    )

    #: 終了タグが来ない要素。深さの計算から除外する。
    VOID_TAGS = {
        "area", "base", "br", "col", "embed", "hr", "img", "input",
        "link", "meta", "param", "source", "track", "wbr",
    }

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        #: 文書全体から拾った行。本文の入れ物が見つからなかったときに使う。
        self.lines: list[str] = []
        #: 本文の入れ物の内側で拾った行。
        self.content_lines: list[str] = []
        self.found_content = False
        self.dropped_chrome = 0

        self._skip_depth = 0
        self._depth = 0
        self._content_depth: int | None = None
        self._buffer: list[str] = []
        self._current_tag: str | None = None
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._list_depth = 0

    # -- ヘルパ --------------------------------------------------------
    def _flush(self) -> str:
        text = " ".join("".join(self._buffer).split())
        self._buffer = []
        return text

    def _emit(self, *lines: str) -> None:
        self.lines.extend(lines)
        if self._content_depth is not None:
            self.content_lines.extend(lines)

    @staticmethod
    def _attr(attrs: list[tuple[str, str | None]], name: str) -> str:
        for key, value in attrs:
            if key == name and value:
                return value.lower()
        return ""

    def _is_chrome(self, attrs: list[tuple[str, str | None]]) -> bool:
        marker = f"{self._attr(attrs, 'id')} {self._attr(attrs, 'class')}"
        return any(m in marker for m in self.CHROME_MARKERS)

    def _is_content_root(self, attrs: list[tuple[str, str | None]]) -> bool:
        element_id = self._attr(attrs, "id")
        classes = set(self._attr(attrs, "class").split())
        return element_id in self.CONTENT_IDS or bool(
            classes & self.CONTENT_CLASSES
        )

    # -- HTMLParser ----------------------------------------------------
    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self.VOID_TAGS:
            # 終了タグが来ないので深さは動かさない。改行だけ拾う。
            if not self._skip_depth and tag == "br":
                self._buffer.append(" ")
            return

        self._depth += 1

        # 既に読み飛ばし中なら、入れ子を数えるだけ。
        if self._skip_depth:
            self._skip_depth += 1
            return

        if tag in self.SKIP_TAGS or self._is_chrome(attrs):
            self._skip_depth = 1
            self.dropped_chrome += 1
            return

        if self._content_depth is None and self._is_content_root(attrs):
            self._content_depth = self._depth
            self.found_content = True

        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in ("td", "th"):
            self._buffer = []
        elif tag in self.HEADINGS or tag == "p":
            self._buffer = []
            self._current_tag = tag
        elif tag == "li":
            self._buffer = []
            self._current_tag = "li"
        elif tag in ("ul", "ol"):
            self._list_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in self.VOID_TAGS:
            return

        if self._skip_depth:
            self._skip_depth -= 1
            self._depth = max(0, self._depth - 1)
            return

        if tag in ("td", "th") and self._row is not None:
            self._row.append(_escape_cell(self._flush()))
        elif tag == "tr" and self._table is not None and self._row is not None:
            self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            trimmed = _trim_empty(self._table)
            if trimmed:
                self._emit("", *_render_table(trimmed), "")
            self._table = None
        elif tag in self.HEADINGS:
            text = self._flush()
            if text:
                level = min(int(tag[1]) + 1, 6)  # 取り込み先の見出し階層に合わせる
                self._emit("", "#" * level + f" {text}", "")
            self._current_tag = None
        elif tag == "p":
            text = self._flush()
            if text:
                self._emit(text, "")
            self._current_tag = None
        elif tag == "li":
            text = self._flush()
            if text:
                indent = "  " * max(0, self._list_depth - 1)
                self._emit(f"{indent}- {text}")
            self._current_tag = None
        elif tag in ("ul", "ol"):
            self._list_depth = max(0, self._list_depth - 1)
            self._emit("")

        # 本文の入れ物を抜けたら、以降は本文ではない。
        if self._content_depth is not None and self._depth == self._content_depth:
            self._content_depth = None
        self._depth = max(0, self._depth - 1)

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if self._current_tag or self._row is not None:
            self._buffer.append(data)


def _read_html(path: Path) -> tuple[list[str], list[str]]:
    parser = _HtmlToMarkdown()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    parser.close()

    warnings: list[str] = []

    # 本文の入れ物が見つかればその中だけを採る。見つからなければ文書全体。
    if parser.found_content and any(line.strip() for line in parser.content_lines):
        source_lines = parser.content_lines
    else:
        source_lines = parser.lines
        if not parser.found_content:
            warnings.append(
                "本文の領域を特定できなかったため、ページ全体を取り込みました。"
                "画面部品が混ざっていないか確認してください。"
            )

    # 連続する空行をまとめる。
    lines: list[str] = []
    for line in source_lines:
        if not line and lines and not lines[-1]:
            continue
        lines.append(line)

    if not any(line.strip() for line in lines):
        warnings.append("本文を抽出できませんでした。HTML の構造を確認してください。")
    return lines, warnings


# ----------------------------------------------------------------------
# CSV / テキスト
# ----------------------------------------------------------------------
def _read_csv(path: Path) -> tuple[list[str], list[str]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    rows = [
        [_escape_cell(cell) for cell in row]
        for row in csv_module.reader(io.StringIO(text))
    ]
    trimmed = _trim_empty(rows)
    if not trimmed:
        return [], ["内容が空でした。"]
    return _render_table(trimmed), []


def _read_text(path: Path) -> tuple[list[str], list[str]]:
    return path.read_text(encoding="utf-8", errors="replace").splitlines(), []


# ----------------------------------------------------------------------
def ingest(
    source: Path, output: Path, *, sheets: list[str] | None = None
) -> IngestResult:
    """設計書を後続工程が読める Markdown に変換する。"""
    if not source.is_file():
        raise IngestError(f"入力ファイルが見つかりません: {source}")

    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise IngestError(
            f"未対応の形式です: {suffix}\n"
            f"対応: {', '.join(sorted(SUPPORTED_SUFFIXES))}\n"
            "PDF は表のレイアウトが失われるため、元の Excel か、"
            "Confluence の HTML エクスポートを使ってください。"
        )

    if suffix in (".xlsx", ".xlsm"):
        body, warnings = _read_xlsx(source, sheets)
    elif suffix in (".html", ".htm"):
        body, warnings = _read_html(source)
    elif suffix == ".csv":
        body, warnings = _read_csv(source)
    else:
        body, warnings = _read_text(source)

    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    header = [
        f"# {source.stem}",
        "",
        "<!--",
        f"  {source.name} から自動変換した設計書。",
        f"  変換日時: {stamp}",
        "",
        "  この文書は体裁を整えず、元の表をそのまま写している。",
        "  読み取れない箇所があれば推測で補わず、open_questions に",
        "  「元の設計書に何を追記してほしいか」の形で挙げること。",
        "-->",
        "",
    ]

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join([*header, *body]) + "\n", encoding="utf-8")

    sections = sum(1 for line in body if line.startswith("## "))
    return IngestResult(
        output=output, source=source, sections=sections, warnings=warnings
    )
